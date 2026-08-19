"""Генератор примера Excel-выгрузки: python scripts/make_example_telemetry.py"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "telemetry" / "ПРИМЕР КНС-54 НА-1 (синтетические данные).xlsx"

START = datetime(2025, 10, 1)
DAYS = 300

HEADERS = [
    "Дата",
    "Давление на приёме, МПа",
    "Давление на выкиде, МПа",
    "Мощность, кВт",
    "Энергопотребление, кВт·ч",
    "Расход суточный, м³",
    "Наработка, ч",
]

P_IN = 1.2
P_OUT = 10.4
ETA_UNIT = 0.62


def rows():
    for day in range(DAYS):
        q_day = 2400.0 + 4.0 * day
        q_hour = q_day / 24.0
        p_hydraulic = q_hour * (P_OUT - P_IN) / 3.6
        power = p_hydraulic / ETA_UNIT
        yield [
            START + timedelta(days=day),
            P_IN,
            P_OUT,
            round(power, 1),
            round(power * 24.0, 1),
            round(q_day, 1),
            24.0,
        ]


def main() -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "НА-1"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows():
        sheet.append(row)
    for index, header in enumerate(HEADERS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(12, len(header) + 2)
    for cell in sheet["A"][1:]:
        cell.number_format = "DD.MM.YYYY"
    sheet.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    book.save(OUT)
    print(f"готово: {OUT.relative_to(ROOT)} — строк {DAYS}")


if __name__ == "__main__":
    main()
