import json
from datetime import datetime

import openpyxl
import pytest

from ppd_audit.db import AuditDatabase
from ppd_audit.db_import import import_excel_telemetry, import_test_telemetry


def test_importer_converts_kns10_pressure_from_kgf_per_cm2(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("kns10bn", "КНС-10 БН", "Бавлынефть", "пластовая", "кнс")
    database.upsert_aggregate("kns10bn", "НА-1", "работа")
    draft = {
        "telemetry": [
            {
                "timestamp": "2026-07-24T00:00:00",
                "metric": "pressure",
                "label": "Давление на приёме",
                "value": 10.0,
                "quality": 0,
            },
            {
                "timestamp": "2026-07-24T00:00:00",
                "metric": "pressure",
                "label": "Давление на выкиде",
                "value": 130.0,
                "quality": 0,
            },
        ]
    }
    path = tmp_path / "КНС-10 БН давления НА-1 из вомбат.xls.json"
    path.write_text(json.dumps(draft), encoding="utf-8")

    stats = import_test_telemetry(database, tmp_path)
    records = database.measurements("kns10bn", "НА-1")

    assert stats.stored == 2
    assert [row["metric"] for row in records] == ["p_in", "p_out"]
    assert records[0]["value"] == pytest.approx(10.0 * 0.0980665)
    assert records[0]["unit"] == "МПа"
    assert records[0]["source_kind"] == "json_draft"
    assert records[0]["source_file"] == path.name
    assert records[0]["source_label"] == "Давление на приёме"


def test_importer_converts_kns97_pressure_from_kgf_per_cm2(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("kns97pren", "КНС-97", "Елховнефть", "пресная", "кнс")
    database.upsert_aggregate("kns97pren", "НА-1", "работа")
    (tmp_path / "КНС-97 ЕН давления НА-1 из вомбат.xls.json").write_text(
        json.dumps(
            {
                "telemetry": [
                    {
                        "timestamp": "2026-07-24T00:00:00",
                        "metric": "pressure",
                        "label": "Давление на выкиде",
                        "value": 100.0,
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    import_test_telemetry(database, tmp_path)

    assert database.measurements("kns97pren", "НА-1")[0]["value"] == pytest.approx(9.80665)


def test_importer_keeps_low_kns97_pressure_values_in_mpa(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("kns97pren", "КНС-97", "Елховнефть", "пресная", "кнс")
    database.upsert_aggregate("kns97pren", "НА-2 ПР", "работа")
    (tmp_path / "КНС-97 ПР ЕН давления НА-02 из вомбат.xls.json").write_text(
        json.dumps(
            {
                "telemetry": [
                    {
                        "timestamp": "2026-07-24T00:00:00",
                        "metric": "pressure",
                        "label": "Давление на выкиде",
                        "value": 11.807,
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    import_test_telemetry(database, tmp_path)

    assert database.measurements("kns97pren", "НА-2 ПР")[0]["value"] == pytest.approx(11.807)


def test_importer_uses_kns10_daily_file_tag_as_aggregate_scope(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("kns10bn", "КНС-10 БН", "Бавлынефть", "пластовая", "кнс")
    database.upsert_aggregate("kns10bn", "НА-1", "работа")
    draft = {
        "telemetry": [
            {
                "timestamp": "2026-07-24T00:00:00",
                "tag": "НА 1",
                "metric": "flow",
                "label": "Расход (м3)",
                "value": 500.0,
            }
        ]
    }
    path = tmp_path / "КНС-10 БН проток за год.xls.json"
    path.write_text(json.dumps(draft), encoding="utf-8")

    import_test_telemetry(database, tmp_path)

    assert database.measurements("kns10bn", "НА-1")[0]["metric"] == "q_day"


def test_importer_preserves_zero_motohours_as_valid_telemetry(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("kns54an", "КНС-54", "Азнакаевскнефть", "агрессивная", "кнс")
    database.upsert_aggregate("kns54an", "НА-1", "работа")
    (tmp_path / "КНС-54 Азн проток за год.xls.json").write_text(
        json.dumps(
            {
                "telemetry": [
                    {
                        "timestamp": "2025-05-15T00:00:00",
                        "tag": "НА 1",
                        "metric": "runtime_h",
                        "label": "Моточасы (мин)",
                        "value": 0.0,
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    stats = import_test_telemetry(database, tmp_path)

    assert stats.stored == 1
    assert database.measurements("kns54an", "НА-1")[0]["value"] == 0.0


def test_importer_keeps_kns97_aggregates_in_one_object(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("kns97pren", "КНС-97/2", "Елховнефть", "пресная", "кнс")
    draft = {
        "telemetry": [
            {
                "timestamp": "2026-07-24T00:00:00",
                "metric": "power_kw",
                "label": "Активная мощность",
                "value": 200.0,
            }
        ]
    }
    path = tmp_path / "КНС-97 ЕН мощность НА-03 из вомбат.xls.json"
    path.write_text(json.dumps(draft), encoding="utf-8")

    import_test_telemetry(database, tmp_path)

    records = database.measurements("kns97pren", "НА-3")
    assert records[0]["metric"] == "power"


def test_importer_maps_kns97_pr_daily_tag_to_na02(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("kns97pren", "КНС-97/2", "Елховнефть", "пресная", "кнс")
    database.upsert_aggregate("kns97pren", "НА-2 ПР", "работа")
    draft = {
        "telemetry": [
            {
                "timestamp": "2026-07-24T00:00:00",
                "tag": "НА 02(ЦНС)",
                "metric": "flow",
                "label": "Расход (м3)",
                "value": 500.0,
            }
        ]
    }
    path = tmp_path / "КНС-97 ПР ЕН проток за год.xls.json"
    path.write_text(json.dumps(draft), encoding="utf-8")

    import_test_telemetry(database, tmp_path)

    assert database.measurements("kns97pren", "НА-2 ПР")[0]["metric"] == "q_day"


def test_importer_maps_knsopu_daily_values_to_aggregate_scope(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("knsopu", "КНС-ОПУ", "Джалильнефть", "агрессивная", "кнс")
    database.upsert_aggregate("knsopu", "НА-1", "работа")
    database.upsert_aggregate("knsopu", "НА-2", "работа")
    draft = {
        "telemetry": [
            {
                "timestamp": "2026-07-24T00:00:00",
                "tag": "НА 1",
                "metric": "flow",
                "label": "Расход (м3)",
                "value": 500.0,
            },
            {
                "timestamp": "2026-07-24T00:00:00",
                "tag": "НА 2",
                "metric": "flow",
                "label": "Расход (м3)",
                "value": 600.0,
            },
        ]
    }
    path = tmp_path / "КНС ОПУ ДжН проток за год.xls.json"
    path.write_text(json.dumps(draft), encoding="utf-8")

    import_test_telemetry(database, tmp_path)

    assert database.measurements("knsopu", "НА-1")[0]["metric"] == "q_day"
    assert database.measurements("knsopu", "НА-2")[0]["value"] == pytest.approx(600.0)


def test_importer_converts_knsopu_pressure_from_kgf_per_cm2(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("knsopu", "КНС-ОПУ", "Джалильнефть", "агрессивная", "кнс")
    database.upsert_aggregate("knsopu", "НА-1", "работа")
    (tmp_path / "КНС-ОПУ ДжН давление НА-1 из вомбат.xls.json").write_text(
        json.dumps(
            {
                "telemetry": [
                    {
                        "timestamp": "2026-07-24T00:00:00",
                        "metric": "pressure",
                        "label": "Давление на выкиде",
                        "value": 130.0,
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    import_test_telemetry(database, tmp_path)

    assert database.measurements("knsopu", "НА-1")[0]["value"] == pytest.approx(
        130.0 * 0.0980665
    )


def test_importer_converts_kns54_kgf_pressure_and_keeps_bg_station_scope(tmp_path):
    database = AuditDatabase(tmp_path / "audit.sqlite")
    database.migrate()
    database.upsert_plant("kns54an", "КНС-54", "Азнакаевскнефть", "пластовая", "кнс")
    database.upsert_aggregate("kns54an", "НА-1", "работа")

    pressures = openpyxl.Workbook()
    pressure_sheet = pressures.active
    pressure_sheet.append(["Дата", "Давление на приёме", "Давление на выкиде"])
    pressure_sheet.append([datetime(2026, 4, 24), 20.0, 155.0])
    pressures.save(tmp_path / "КНС-54 НА-1 давления.xlsx")

    bg = openpyxl.Workbook()
    bg_sheet = bg.active
    bg_sheet.append(["Дата", "Давление в коллекторе"])
    bg_sheet.append([datetime(2026, 4, 24), 126.0])
    bg.save(tmp_path / "КНС-54 давление на БГ.xlsx")

    energy = openpyxl.Workbook()
    energy_sheet = energy.active
    energy_sheet.append(["Дата", "Активная мощность"])
    energy_sheet.append([datetime(2026, 4, 24), 0.0])
    energy.save(tmp_path / "КНС-54 НА-1 энергия.xlsx")

    stats = import_excel_telemetry(database, tmp_path)
    records = database.measurements("kns54an", "НА-1")
    station = database.measurements("kns54an")

    assert stats.stored == 4
    assert records[0]["value"] == pytest.approx(20.0 * 0.0980665)
    assert records[1]["value"] == pytest.approx(155.0 * 0.0980665)
    assert records[2]["metric"] == "power"
    assert records[2]["value"] == 0.0
    assert any(row["metric"] == "p_bg" for row in station)
