"""Семантический парсер инженерных файлов «… расчет.xlsx».

Файлы — ручные расчёты энергоаудита: содержат и ВХОДЫ (паспорт насоса/ЭД, режим),
и ЭТАЛОННЫЕ ВЫХОДЫ (УРЭ, КПД, напор должный, годовые потери). Шаблоны у объектов
различаются (метки в колонке D или F; значения в разных колонках) — поэтому поиск
идёт ПО ТЕКСТУ МЕТОК, а значения берутся из колонок реальных агрегатов.

Возвращает ObjectSpec с заполненными regime и reference у каждого агрегата.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from ..spec import (
    AggregateSpec,
    Branch,
    MotorSpec,
    ObjectSpec,
    PumpSpec,
    ReferenceOutputs,
    RegimeMeasurement,
    WaterType,
    infer_motor_synchronous,
    infer_pump_kind,
)


@dataclass
class CellBinding:
    object_id: str
    object_name: str
    water_type: str
    aggregate_id: str
    role: str
    field: str
    label: str
    method_ref: str
    unit: str
    sheet: str
    label_cell: str | None
    value_cell: str | None
    raw: Any
    value: float | None


@dataclass
class ParsedCalc:
    spec: ObjectSpec
    cells: list[CellBinding]


@dataclass
class _CellSource:
    field: str
    sheet: str
    label_cell: str | None
    value_cell: str | None
    raw: Any
    value: float | None


_FIELD_META = {
    "pump.model": ("Модель насоса", "паспорт", ""),
    "pump.h_nom": ("H номинальный насоса", "29", "м"),
    "pump.q_nom": ("Q номинальный насоса", "паспорт", "м³/ч"),
    "pump.power_nom": ("P потребляемая номинальная насоса", "паспорт", "кВт"),
    "pump.n_rpm": ("Обороты насоса", "паспорт", "об/мин"),
    "pump.eta_nom": ("КПД насоса номинальный", "14/15", "о.е."),
    "pump.p_motor_rec": ("P рекомендованного ЭД", "паспорт", "кВт"),
    "motor.model": ("Модель ЭД", "паспорт", ""),
    "motor.voltage_kv": ("U ЭД", "паспорт", "кВ"),
    "motor.i_nom": ("I номинальный ЭД", "паспорт", "А"),
    "motor.p_nom": ("P номинальная ЭД", "24", "кВт"),
    "motor.n_rpm": ("Обороты ЭД", "паспорт", "об/мин"),
    "motor.eta_nom": ("КПД ЭД номинальный", "14/15,24", "о.е."),
    "motor.cos_phi": ("cos φ ЭД", "паспорт", ""),
    "t": ("T, время работы", "7/12/16", "ч"),
    "q_day": ("Qсут, суточная перекачка", "7/16", "м³/сут"),
    "w": ("W, суточный расход ЭЭ", "12/16", "кВт·ч/сут"),
    "p_electric": ("Pэл, электрическая мощность", "12/13/24", "кВт"),
    "rho": ("ρ, плотность", "8", "кг/м³"),
    "p_in": ("pвх, давление на входе", "8/11/17", "МПа"),
    "p_out": ("pвых, давление на выходе", "8/11/17", "МПа"),
    "q_fact": ("Q, фактическая производительность", "7/11/13", "м³/ч"),
    "p_bg": ("pБГ, давление на БГ", "45", "МПа"),
    "t_year": ("Tгод, годовая наработка", "44/45", "ч/год"),
    "h_fact": ("Hф, напор факт", "8", "м"),
    "h_due": ("Hд, напор должный", "29", "м"),
    "p_hydraulic": ("Pгидр, гидравлическая мощность", "11", "кВт"),
    "eta_fact": ("КПД НА факт", "13", "о.е."),
    "eta_nom": ("КПД НА номинальный", "14/15", "о.е."),
    "sec_fact": ("УРЭ факт", "16", "кВт·ч/м³"),
    "sec_calc": ("УРЭ расчётный", "17", "кВт·ч/м³"),
    "load_factor": ("K загрузки ЭД", "24", "о.е."),
    "dw_efficiency": ("ΔW КПД", "44", "кВт·ч/год"),
    "dw_throttle": ("ΔW дрос", "45", "кВт·ч/год"),
}


def _addr(r: int | None, c: int | None) -> str | None:
    return f"{get_column_letter(c)}{r}" if r and c else None


def _safe_raw(v):
    if v is None or isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def _norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip().lower().replace("ё", "е"))


def _num(v) -> float | None:
    """Число или None (мусор #REF!/#DIV/0!/текст → None)."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".") if v is not None else ""
    try:
        return float(s)
    except ValueError, TypeError:
        return None


class _Sheet:
    """2D-сетка значений листа (1-индексация) с поиском по меткам."""

    def __init__(self, rows: list[tuple], name: str = ""):
        self.name = name
        self.rows = rows
        self.nrow = len(rows)
        self.ncol = max((len(r) for r in rows), default=0)

    def cell(self, r: int, c: int):
        if 1 <= r <= self.nrow and 1 <= c <= len(self.rows[r - 1]):
            return self.rows[r - 1][c - 1]
        return None

    def find(
        self,
        *patterns: str,
        exclude: tuple[str, ...] = (),
        col: int | None = None,
    ) -> tuple[int, int] | None:
        """Первая ячейка, чей текст содержит любой pattern и не содержит exclude.

        Если задан col — поиск только в этой колонке (устойчивость к меткам-двойникам).
        """
        cols_range = [col] if col else None
        for r in range(1, self.nrow + 1):
            rng = cols_range or range(1, len(self.rows[r - 1]) + 1)
            for c in rng:
                t = _norm(self.cell(r, c))
                if not t:
                    continue
                if any(p in t for p in patterns) and not any(e in t for e in exclude):
                    return (r, c)
        return None

    def value_at_cols(self, r: int, cols: list[int]) -> list[float | None]:
        return [_num(self.cell(r, c)) for c in cols]


def _pick_sheet(wb) -> _Sheet | None:
    """Выбрать лист с расчётом (содержит метку «УРЭ факт»)."""
    best = None
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        sh = _Sheet(rows, name)
        if sh.find("урэ факт"):
            return sh
        if best is None and sh.nrow > 5:
            best = sh
    return best


def _aggregate_columns(sh: _Sheet) -> list[int]:
    """Колонки реальных агрегатов: где и Qсут, и рвых, и рвх — положительные числа."""

    def numeric_cols(anchor: tuple[int, int] | None) -> set[int]:
        if not anchor:
            return set()
        r = anchor[0]
        return {
            c
            for c in range(anchor[1] + 1, len(sh.rows[r - 1]) + 1)
            if (_num(sh.cell(r, c)) or 0) > 0
        }

    qsut = sh.find("суточная перекачка")
    pout = sh.find("давление на выходе", "давление на вых")
    pin = sh.find("давление на входе", "давление на вх")
    cols = numeric_cols(qsut) & numeric_cols(pout) & numeric_cols(pin)
    # отсечь явный мусор: оставить не более 4 колонок подряд
    return sorted(cols)[:4]


def _passport_block(
    sh: _Sheet,
    header_patterns: tuple[str, ...],
    field_map: dict[str, tuple[str, ...]],
    n: int,
) -> list[dict]:
    """Распарсить блок паспорта (насоса/ЭД): по заголовку → карта колонок → n строк."""
    hdr = sh.find(*header_patterns)
    if not hdr:
        return [{} for _ in range(n)]
    r0, c0 = hdr
    # карта поле→колонка по заголовочной строке
    colmap: dict[str, int] = {}
    name_col = c0  # столбец с типом/моделью
    for c in range(c0, len(sh.rows[r0 - 1]) + 1):
        t = _norm(sh.cell(r0, c))
        for field, pats in field_map.items():
            if field not in colmap and any(p in t for p in pats):
                colmap[field] = c
    # строки агрегатов идут ниже заголовка; берём те, где есть число в любой числовой колонке
    out: list[dict] = []
    r = r0 + 1
    while r <= sh.nrow and len(out) < n:
        has_num = any(_num(sh.cell(r, c)) is not None for c in colmap.values())
        model = sh.cell(r, name_col)
        if has_num and model is not None:
            rec: dict[str, Any] = {"model": str(model).strip()}
            cells = {
                "model": _CellSource(
                    "model",
                    sh.name,
                    _addr(r0, name_col),
                    _addr(r, name_col),
                    _safe_raw(model),
                    None,
                )
            }
            for field, c in colmap.items():
                raw = sh.cell(r, c)
                rec[field] = _num(raw)
                cells[field] = _CellSource(
                    field, sh.name, _addr(r0, c), _addr(r, c), _safe_raw(raw), rec[field]
                )
            rec["__cells__"] = cells
            out.append(rec)
        elif out:  # блок закончился
            break
        r += 1
    while len(out) < n:
        out.append({})
    return out


# Канонические метки строк режима/результатов: поле → (паттерны, исключения).
# Значение читается на колонках агрегатов.
_REGIME_LABELS = {
    "t": (("время работы на",), ()),
    "q_day": (("суточная перекачка",), ()),
    "w": (("суточный расход ээ", "расход эл. энергии", "расход ээ"), ()),
    "p_electric": (("электрическая мощность",), ()),
    "rho": (("плотность",), ()),
    "p_in": (("давление на входе", "давление на вх"), ()),
    "p_out": (("давление на выходе", "давление на вых"), ()),
    "q_fact": (("фактич произв",), ()),
    "p_bg": (("давление на бг",), ()),
    "t_year": (("годовая нараб",), ()),
}
_REF_LABELS = {
    "h_fact": (("напор",), ("должн", "снижен", "на в сут")),
    "h_due": (("напор должн",), ()),
    "p_hydraulic": (("гидравлическая мощность",), ("на бг",)),
    "eta_fact": (("кпд на среднесут",), ()),
    "eta_nom": (("кпд на в ном",), ()),
    "sec_fact": (("урэ факт",), ()),
    "sec_calc": (("урэ расч",), ()),
    "load_factor": (("коэф загруз",), ()),
    "dw_efficiency": (("wкпд",), ()),
    "dw_throttle": (("wдрос", "wдросс"), ()),
}


def _regime_label_col(sh: _Sheet) -> int | None:
    """Колонка меток режима = колонка метки «Суточная перекачка»."""
    anchor = sh.find("суточная перекачка")
    return anchor[1] if anchor else None


def _read_label_values(
    sh: _Sheet,
    labels: dict,
    cols: list[int],
    label_col: int | None,
) -> dict[str, list]:
    sources = _read_label_sources(sh, labels, cols, label_col)
    return {field: [src.value for src in cells] for field, cells in sources.items()}


def _read_label_sources(
    sh: _Sheet,
    labels: dict,
    cols: list[int],
    label_col: int | None,
) -> dict[str, list[_CellSource]]:
    out: dict[str, list] = {}
    for field, (pats, excl) in labels.items():
        anchor = sh.find(*pats, exclude=excl, col=label_col)
        if anchor is None:  # запас: глобальный поиск
            anchor = sh.find(*pats, exclude=excl)
        if anchor:
            out[field] = [
                _CellSource(
                    field,
                    sh.name,
                    _addr(*anchor),
                    _addr(anchor[0], col),
                    _safe_raw(sh.cell(anchor[0], col)),
                    _num(sh.cell(anchor[0], col)),
                )
                for col in cols
            ]
        else:
            out[field] = [_CellSource(field, sh.name, None, None, None, None) for _ in cols]
    return out


_UNIT_SCALE = {"тыс. квт": 1000.0, "тыс квт": 1000.0}


def _dw_scale(sh: _Sheet, label_pats: tuple[str, ...]) -> float:
    """Множитель единиц для годовых потерь (тыс. кВт·ч → ×1000)."""
    anchor = sh.find(*label_pats)
    if not anchor:
        return 1.0
    r, c = anchor
    for cc in range(c, min(c + 4, len(sh.rows[r - 1]) + 1)):
        t = _norm(sh.cell(r, cc))
        for key, scale in _UNIT_SCALE.items():
            if key in t:
                return scale
    return 1.0


_WATER_BY_PATH = {
    "пресная": WaterType.fresh,
    "агрессив": WaterType.aggressive,
    "пластов": WaterType.formation,
}


def _water_type(path: Path) -> WaterType:
    p = _norm(str(path))
    for key, wt in _WATER_BY_PATH.items():
        if key in p:
            return wt
    return WaterType.fresh


def _bind_cell(
    object_id: str,
    object_name: str,
    water_type: str,
    aggregate_id: str,
    role: str,
    field: str,
    src: _CellSource,
    value: float | None = None,
) -> CellBinding:
    label, method_ref, unit = _FIELD_META.get(field, (field, "", ""))
    return CellBinding(
        object_id=object_id,
        object_name=object_name,
        water_type=water_type,
        aggregate_id=aggregate_id,
        role=role,
        field=field,
        label=label,
        method_ref=method_ref,
        unit=unit,
        sheet=src.sheet,
        label_cell=src.label_cell,
        value_cell=src.value_cell,
        raw=src.raw,
        value=src.value if value is None else value,
    )


def parse_calc_file(path: Path, object_id: str, object_name: str) -> ObjectSpec:
    """Распарсить «… расчет.xlsx» в ObjectSpec (входы + эталонные выходы)."""
    return parse_calc_file_with_cells(path, object_id, object_name).spec


def parse_calc_file_with_cells(path: Path, object_id: str, object_name: str) -> ParsedCalc:
    """Распарсить «… расчет.xlsx» и вернуть ObjectSpec плюс привязку полей к ячейкам."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sh = _pick_sheet(wb)
    finally:
        wb.close()
    if sh is None:
        raise ValueError(f"не найден лист расчёта в {path}")

    cols = _aggregate_columns(sh)
    if not cols:
        raise ValueError(f"не определены колонки агрегатов в {path}")

    pumps = _passport_block(
        sh,
        ("тип насоса",),
        {
            "h_nom": ("нном",),
            "q_nom": ("qном",),
            "power_nom": ("pпотреб", "потреб"),
            "n_rpm": ("об/мин",),
            "eta_nom": ("кпд",),
            "p_motor_rec": ("рекоменд",),
        },
        len(cols),
    )
    # ограничить число агрегатов количеством паспортных строк насосов (с моделью)
    n_real = sum(1 for p in pumps if p.get("model"))
    if n_real:
        cols = cols[:n_real]
        pumps = pumps[: len(cols)]
    motors = _passport_block(
        sh,
        ("тип эл",),
        {
            "voltage_kv": ("u, кв", "u,кв"),
            "i_nom": ("iном",),
            "p_nom": ("pном",),
            "n_rpm": ("об/мин",),
            "eta_nom": ("кпд",),
            "cos_phi": ("cos",),
        },
        len(cols),
    )

    label_col = _regime_label_col(sh)
    regime_sources = _read_label_sources(sh, _REGIME_LABELS, cols, label_col)
    ref_sources = _read_label_sources(sh, _REF_LABELS, cols, label_col)
    regime = {field: [src.value for src in cells] for field, cells in regime_sources.items()}
    ref = {field: [src.value for src in cells] for field, cells in ref_sources.items()}
    dw_scale = _dw_scale(sh, ("wкпд",))

    water = _water_type(path)
    aggregates: list[AggregateSpec] = []
    cells: list[CellBinding] = []
    for i, col in enumerate(cols):
        p = pumps[i] if i < len(pumps) else {}
        m = motors[i] if i < len(motors) else {}
        pump_model = p.get("model", "")
        pump = PumpSpec(
            model=pump_model,
            kind=infer_pump_kind(pump_model, p.get("n_rpm")),
            q_nom=p.get("q_nom"),
            h_nom=p.get("h_nom"),
            eta_nom=p.get("eta_nom"),
            power_nom=p.get("power_nom"),
            n_rpm=p.get("n_rpm"),
        )
        motor = MotorSpec(
            model=m.get("model", ""),
            synchronous=infer_motor_synchronous(m.get("model", "")),
            p_nom=m.get("p_nom"),
            eta_nom=m.get("eta_nom"),
            cos_phi=m.get("cos_phi"),
            voltage_kv=m.get("voltage_kv"),
            i_nom=m.get("i_nom"),
            n_rpm=m.get("n_rpm"),
        )

        rho = regime["rho"][i] or 1000.0
        p_in, p_out = regime["p_in"][i], regime["p_out"][i]
        if p_in is None or p_out is None:
            continue
        # Правдоподобие: давления ППД — единицы МПа (макс ~16 по ограничениям).
        # Значения в сотни/тысячи — захват чужой колонки (напор в м, проектные
        # значения) → колонка не является реальным агрегатом, пропускаем.
        if p_out > 30.0 or p_in > 30.0:
            continue
        # Ноль в ячейках расхода/энергии/наработки — «не заполнено» (черновики
        # инженеров, напр. W=0 у КНС-85 НА-1), а не физический ноль → None,
        # чтобы ядро выбрало запасной путь (УРЭ_ф = P_эл/Q вместо W/Q_сут).
        _pos = lambda v: v if (v is not None and v > 0) else None  # noqa: E731
        rm = RegimeMeasurement(
            rho=rho,
            p_in=p_in,
            p_out=p_out,
            q_day=_pos(regime["q_day"][i]),
            t=_pos(regime["t"][i]),
            w=_pos(regime["w"][i]),
            p_electric=_pos(regime["p_electric"][i]),
            q_fact=_pos(regime["q_fact"][i]),
            p_bg=_pos(regime["p_bg"][i]),
            t_year=_pos(regime["t_year"][i]),
        )
        ro = ReferenceOutputs(
            h_fact=ref["h_fact"][i],
            h_due=ref["h_due"][i],
            eta_fact=ref["eta_fact"][i],
            eta_nom=ref["eta_nom"][i],
            sec_fact=ref["sec_fact"][i],
            sec_calc=ref["sec_calc"][i],
            load_factor=ref["load_factor"][i],
            p_hydraulic=ref["p_hydraulic"][i],
            dw_efficiency=(dw_efficiency * dw_scale)
            if (dw_efficiency := ref["dw_efficiency"][i]) is not None
            else None,
            dw_throttle=(dw_throttle * dw_scale)
            if (dw_throttle := ref["dw_throttle"][i]) is not None
            else None,
            t_year=regime["t_year"][i],
        )
        aggregate_id = f"НА-{i + 1}"
        for field, src in p.get("__cells__", {}).items():
            cells.append(
                _bind_cell(
                    object_id,
                    object_name,
                    water.value,
                    aggregate_id,
                    "passport",
                    f"pump.{field}",
                    src,
                )
            )
        for field, src in m.get("__cells__", {}).items():
            cells.append(
                _bind_cell(
                    object_id,
                    object_name,
                    water.value,
                    aggregate_id,
                    "passport",
                    f"motor.{field}",
                    src,
                )
            )
        for field, sources in regime_sources.items():
            cells.append(
                _bind_cell(
                    object_id, object_name, water.value, aggregate_id, "input", field, sources[i]
                )
            )
        for field, sources in ref_sources.items():
            value = sources[i].value
            if field in ("dw_efficiency", "dw_throttle") and value is not None:
                value *= dw_scale
            cells.append(
                _bind_cell(
                    object_id,
                    object_name,
                    water.value,
                    aggregate_id,
                    "reference",
                    field,
                    sources[i],
                    value=value,
                )
            )
        aggregates.append(
            AggregateSpec(
                id=aggregate_id, role="работа", pump=pump, motor=motor, regime=rm, reference=ro
            )
        )

    spec = ObjectSpec(
        id=object_id,
        name=object_name,
        water_type=water,
        branch=Branch.kns,
        source=str(path),
        aggregates=aggregates,
    )
    return ParsedCalc(spec=spec, cells=cells)


def apply_t_year_overrides(parsed: ParsedCalc, overrides: dict[str, dict]) -> None:
    """Применить документированные допущения годовой наработки для сверки."""
    for aggregate_id, override in overrides.items():
        aggregate = parsed.spec.aggregate(aggregate_id)
        if aggregate.regime is None:
            raise ValueError(f"у агрегата {aggregate_id} нет измеренного режима")
        value = float(override["value"])
        comment = override["comment"]
        aggregate.regime.t_year = value
        for cell in parsed.cells:
            if (cell.aggregate_id, cell.role, cell.field) != (aggregate_id, "input", "t_year"):
                continue
            cell.value = value
            cell.raw = f"{override['formula']}; {comment}"
            cell.label = f"{cell.label}; {comment}"
