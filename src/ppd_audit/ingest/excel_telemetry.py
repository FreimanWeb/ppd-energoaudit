"""Draft Excel -> telemetry JSON exporter for the NTU archive."""

from __future__ import annotations

import json
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import xlrd

from ..config import project_root


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("ё", "е"))


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and math.isnan(value):
        return None
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)) and not math.isnan(float(value)):
        return float(value)
    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return None


def _timestamp(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).isoformat()
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, fmt).isoformat()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).isoformat()
    except ValueError:
        return None


def _metric(header: str) -> tuple[str, str] | None:
    text = _norm(header)
    if not text or "кач" in text:
        return None
    if "плотн" in text:
        return "density_kg_m3", "кг/м³" if "кг/м3" in text or "кг/м³" in text else ""
    if "мощ" in text:
        return "power_kw", "кВт" if "квт" in text else ""
    if "энерг" in text or "квт" in text:
        return "energy_kwh", "кВт·ч" if "квт" in text else ""
    if "давлен" in text:
        return "pressure", "МПа" if "мпа" in text else ""
    if "уд" in text and "расход" in text:
        return "sec_kwh_per_m3", "кВт·ч/м³"
    if "расход" in text or "перекач" in text or "закач" in text:
        return "flow", "м³" if "м3" in text or "м³" in text else ""
    if "моточас" in text or "время" in text or "нараб" in text:
        return "runtime_h", "ч"
    if "урэ" in text:
        return "sec_kwh_per_m3", "кВт·ч/м³"
    return None


def _read_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            {"name": ws.title, "rows": [list(r) for r in ws.iter_rows(values_only=True)]}
            for ws in wb.worksheets
        ]
    finally:
        wb.close()


def _read_xls(path: Path) -> list[dict]:
    book = xlrd.open_workbook(path, on_demand=True)
    try:
        sheets = []
        for name in book.sheet_names():
            sh = book.sheet_by_name(name)
            rows = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        row.append(xlrd.xldate.xldate_as_datetime(cell.value, book.datemode))
                    else:
                        row.append(cell.value)
                rows.append(row)
            sheets.append({"name": name, "rows": rows})
        return sheets
    finally:
        book.release_resources()


def _table_from_header(sheet_name: str, rows: list[list[Any]], header_index: int) -> dict:
    headers = [str(c).strip() if c is not None else "" for c in rows[header_index]]
    date_cols = [
        i for i, h in enumerate(headers) if _norm(h) in {"дата", "время"} or "дата" in _norm(h)
    ]
    tag_cols = [
        i
        for i, h in enumerate(headers)
        if any(x in _norm(h) for x in ("тех", "место", "тег", "описание"))
    ]
    quality_cols = [i for i, h in enumerate(headers) if "кач" in _norm(h)]
    metrics = []
    for i, header in enumerate(headers):
        parsed_metric = _metric(header)
        if parsed_metric:
            metrics.append((i, *parsed_metric))
    telemetry = []
    for row_number, row in enumerate(rows[header_index + 1 :], header_index + 2):
        ts = None
        for i in date_cols:
            if i < len(row):
                ts = _timestamp(row[i])
                if ts:
                    break
        if not ts:
            continue
        tag = next((str(row[i]).strip() for i in tag_cols if i < len(row) and row[i]), sheet_name)
        quality = next(
            (_json_value(row[i]) for i in quality_cols if i < len(row) and row[i] is not None), None
        )
        for col, metric, unit in metrics:
            value = _number(row[col]) if col < len(row) else None
            if value is None:
                continue
            item = {
                "sheet": sheet_name,
                "row": row_number,
                "timestamp": ts,
                "tag": tag,
                "metric": metric,
                "label": headers[col],
                "unit": unit,
                "value": value,
            }
            if quality is not None:
                item["quality"] = quality
            telemetry.append(item)
    return {"header_row": header_index + 1, "headers": headers, "telemetry": telemetry}


def _timeseries(sheet_name: str, rows: list[list[Any]]) -> list[dict]:
    signal = sheet_name
    out = []
    for row_number, row in enumerate(rows, 1):
        if not row:
            continue
        ts = _timestamp(row[0])
        if not ts:
            if row[0]:
                signal = str(row[0]).strip()
            continue
        value = _number(row[1]) if len(row) > 1 else None
        if value is None:
            continue
        metric, unit = _metric(signal) or ("value", "")
        out.append({
            "sheet": sheet_name,
            "row": row_number,
            "timestamp": ts,
            "tag": sheet_name,
            "metric": metric,
            "label": signal,
            "unit": unit,
            "value": value,
        })
    return out


def build_excel_telemetry(path: Path, source_root: Path | None = None) -> dict:
    source_root = source_root or project_root()
    sheets = _read_xls(path) if path.suffix.lower() == ".xls" else _read_xlsx(path)
    payload_sheets = []
    telemetry = []
    for sheet in sheets:
        rows = sheet["rows"]
        tables = []
        for i, row in enumerate(rows):
            headers = [_norm(c) for c in row]
            has_date = any("дата" in h or h == "время" for h in headers)
            has_metric = any(_metric(str(c)) for c in row)
            if has_date and has_metric:
                table = _table_from_header(sheet["name"], rows, i)
                tables.append({k: v for k, v in table.items() if k != "telemetry"})
                telemetry.extend(table["telemetry"])
                break
        if not tables:
            telemetry.extend(_timeseries(sheet["name"], rows))
        payload_sheets.append({
            "name": sheet["name"],
            "nrows": len(rows),
            "ncols": max((len(r) for r in rows), default=0),
            "tables": tables,
        })
    return {
        "source": {
            "path": str(path),
            "relative_path": str(path.relative_to(source_root)),
            "format": path.suffix.lower().lstrip("."),
        },
        "schema": "draft.excel.telemetry.v1",
        "sheets": payload_sheets,
        "telemetry": telemetry,
    }


def export_excel_telemetry(path: Path, out_root: Path, source_root: Path | None = None) -> Path:
    source_root = source_root or project_root()
    payload = build_excel_telemetry(path, source_root)
    out_path = out_root / path.relative_to(source_root).with_name(path.name + ".json")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def export_all_excel_telemetry(
    source_root: Path | None = None, out_root: Path | None = None
) -> list[Path]:
    root = project_root()
    source_root = source_root or root / "data" / "raw" / "ntu"
    out_root = out_root or root / "data" / "generated" / "excel_telemetry"
    files = sorted([p for ext in ("*.xlsx", "*.xls") for p in source_root.rglob(ext)])
    return [export_excel_telemetry(path, out_root, source_root) for path in files]
